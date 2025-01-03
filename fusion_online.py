from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from datetime import datetime
import os

app = Flask(__name__)

# Définir le chemin de base local
BASE_PATH = os.path.join(os.getcwd(), "data")
FILE_PATHS = {
    "clients": os.path.join(BASE_PATH, "Clients.txt"),
    "createurs": os.path.join(BASE_PATH, "Createur-OS.txt"),
    "sous_traitants": os.path.join(BASE_PATH, "Sous-Traitants.txt"),
    "categories": os.path.join(BASE_PATH, "list_po_issues.txt"),
}

# Lire un fichier en gérant les encodages
def read_file_from_directory(files_path):
    try:
        with open(files_path, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file if line.strip()]
    except UnicodeDecodeError:
        try:
            with open(files_path, 'r', encoding='iso-8859-1') as file:
                return [line.strip() for line in file if line.strip()]
        except Exception as ee:
            print(f"Erreur lors de la lecture du fichier : {files_path}. Détails : {ee}")
            return []
    except FileNotFoundError:
        print(f"Erreur : Fichier introuvable -> {files_path}")
        return []  # Retourne une liste vide si le fichier est introuvable

# Vérifier si le répertoire `data` est accessible et gérer les fichiers manquants
def get_data_from_directory():
    if not os.path.exists(BASE_PATH):
        return {
            "error_message": "Erreur : Le répertoire local 'data' est introuvable. Veuillez vérifier la configuration.",
            "directory_path": BASE_PATH,
            "file_errors": [],
            "clients": [],
            "createurs": [],
            "flux": [],
            "sous_traitants": [],
            "categories": []
        }

    # Si le répertoire est accessible, essayer de lire les fichiers
    file_errors = []
    data = {
        "clients": read_file_from_directory(FILE_PATHS["clients"]),
        "createurs": read_file_from_directory(FILE_PATHS["createurs"]),
        "flux": ['UNS', 'SERV'],
        "sous_traitants": read_file_from_directory(FILE_PATHS["sous_traitants"]),
        "categories": read_file_from_directory(FILE_PATHS["categories"]),
    }

    # Vérifier les fichiers manquants ou vides
    for key, file_path in FILE_PATHS.items():
        if not data[key]:
            file_errors.append(f"Le fichier {os.path.basename(file_path)} est introuvable ou vide.")

    return {
        "error_message": None,
        "directory_path": BASE_PATH,
        "file_errors": file_errors,
        **data,
    }

# Enregistrer les données dans un fichier Excel
def enregistrer_donnees(numero_os, createur, num_po, client, flux, sous_traitant, categorie, site="WRS",
                        fichier="RESULTAT A TELECHARGER EN LOCAL.xlsx"):
    chemin_fichier = os.path.join(os.getcwd(), fichier)

    # Extraire la cause et la description de la catégorie
    if "-" in categorie:
        cause, description = map(str.strip, categorie.split("-", 1))
    else:
        cause, description = categorie, ""

    # Obtenir la date et l'utilisateur
    date_actuelle = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    utilisateur = os.getlogin()

    if not os.path.exists(chemin_fichier):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Date d'ouverture", "QUI", "OS", "CREE PAR", "PO", "CLIENTS",
            "FLUX", "Sous-traitant", "Cause", "description anomalie", "Site"
        ])
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(chemin_fichier)

    wb = load_workbook(chemin_fichier)
    ws = wb.active
    ws.append([
        date_actuelle, utilisateur, numero_os, createur, num_po, client,
        flux, sous_traitant, cause, description, site
    ])

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = max(len(str(ws.cell(row=1, column=col).value)) + 5, 15)

    wb.save(chemin_fichier)
    print(f"Données enregistrées dans '{fichier}' avec succès.")

@app.route('/')
def home():
    datas = get_data_from_directory()

    if datas["error_message"]:
        return render_template(
            'file_error.html',
            error_message=datas["error_message"],
            directory_path=datas["directory_path"]
        )

    if datas["file_errors"]:
        return render_template(
            'file_error.html',
            file_errors=datas["file_errors"],
            directory_contact="raphael.carabeuf@safrangroup.com"
        )

    return render_template(
        'index.html',
        error_message=None,
        directory_path=datas["directory_path"],
        file_errors=datas["file_errors"],
        clients=datas["clients"],
        createurs=datas["createurs"],
        flux=datas["flux"],
        sous_traitants=datas["sous_traitants"],
        categories=datas["categories"]
    )

@app.route('/submit', methods=['POST'])
def submit():
    numero_os = request.form.get('os_number', '').strip()
    createur = request.form.get('creator', '').strip()
    num_po = request.form.get('po_number', '').strip()
    client = request.form.get('client', '').strip()
    flux = request.form.get('flux', '').strip()
    sous_traitant = request.form.get('contractor', '').strip()
    categorie = request.form.get('category', '').strip()

    try:
        enregistrer_donnees(
            numero_os=numero_os,
            createur=createur,
            num_po=num_po,
            client=client,
            flux=flux,
            sous_traitant=sous_traitant,
            categorie=categorie
        )
        message = "Données enregistrées avec succès !"
    except Exception as e:
        message = f"Erreur lors de l'enregistrement des données : {str(e)}"

    return render_template(
        'index.html',
        error_message=None,
        directory_path=None,
        file_errors=[],
        clients=get_data_from_directory()["clients"],
        createurs=get_data_from_directory()["createurs"],
        flux=get_data_from_directory()["flux"],
        sous_traitants=get_data_from_directory()["sous_traitants"],
        categories=get_data_from_directory()["categories"],
        message=message
    )


if __name__ == '__main__':
    app.run(debug=True)
